import openpyxl
import os

def gen(lang):
    # 打开 excel 文档
    workbook_dir = 'languages\\'+lang+'.xlsx'
    try:
        workbook = openpyxl.load_workbook(workbook_dir)
    except FileNotFoundError:
        input(f"{workbook_dir} not found. Press Enter to close...")
        sys.exit()

    # 创建对应语言文件夹，在里面一个新的函数文件
    if os.path.exists(lang):
        print(f'folder {lang} already exists')
        return
    os.makedirs(lang)
    output_file = open(lang+'\\init.mcfunction', 'w', encoding='utf-8')

    # 写入开头信息
    sheet_author_info = workbook['author_info']
    translator = sheet_author_info.cell(row=2, column=3).value
    if translator is None:
        translator = ''
    translator_info = sheet_author_info.cell(row=3, column=3).value
    if translator_info is None:
        translator_info = ''
    output_file.write('# language: ' + lang + '\n')
    output_file.write('# translator: ' + translator + '\n')
    output_file.write('# Contact information/website: ' + translator_info + '\n\n')
    output_file.write('# ================ common translation ================' + '\n\n')
    
    # 以下处理通用翻译键
    # 打开通用翻译键的表
    sheet_common = workbook['common']
    
    max_row = sheet_common.max_row # 最大行数

    # 从第 2 行开始遍历每行
    for i in range(2,max_row+1):
        # 获取第 B 列也就是第 2 列的值
        cell = sheet_common.cell(row=i, column=2)
        value = cell.value
        # 如果当前单元格是空的，或者它的值不是以 "g_lang" 开头，跳过
        if value is None or not value.startswith("g_lang"):
            continue;
        # 到这里找到一个翻译键，检查其右侧格子有没有通用键
        # 同行第 7 列即为值，输出命令
        translation = sheet_common.cell(row=i, column=7).value
        if translation is None:
            output_file.write(f"common translation error in row {i} col 7\n")
            continue
        output_file.write(f'data modify storage global_shop:storage g_lang."{value[7:]}" set value "{translation}"\n')

    output_file.write('\n# ================ common translation ================' + '\n\n')

###########################################################################################

    output_file.write('# ================ translation ================' + '\n\n')

    # 以下处理独立翻译键
    # 打开独立翻译键的表
    sheet = workbook['translation']
    # 获取工作表的大小
    max_row = sheet.max_row # 最大行数

    # 从第 3 行开始遍历每行
    for i in range(3,max_row+1):
        # 获取第 B 列也就是第 2 列的值
        cell = sheet.cell(row=i, column=2)
        value = cell.value
        # 如果当前单元格是空的，或者它的值不是以 "g_lang" 开头，跳过
        if value is None or not value.startswith("g_lang"):
            continue;
        # 到这里找到一个翻译键，检查其右侧格子有没有通用键
        right_cell = sheet.cell(row=i, column=3)
        right_value = right_cell.value
        if right_value is None:
            # 右侧格子没有通用键，说明这是个独立翻译键
            # 同行第 8 列即为值，输出命令
            translation = sheet.cell(row=i, column=8).value
            if translation is None:
                output_file.write(f"translation error in row {i} col 8\n")
                continue;
            output_file.write(f'data modify storage global_shop:storage g_lang."{value[7:]}" set value "{translation}"\n')
        else:
            # 引用了通用翻译键
            output_file.write(f'data modify storage global_shop:storage g_lang."{value[7:]}" set from storage global_shop:storage g_lang."{right_value[7:]}"\n')

    output_file.write('\n# ================ translation ================')

    output_file.write('\nreturn 1')

    # 关闭文件
    workbook.close()
    output_file.close()

    print(f'gen mcf done, please check the folder "{lang}"')
    input(f"Press Enter to close...")

lang = input("Language translation workbook name: ")
gen(lang)
